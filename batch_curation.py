## By Anqi Lin 08/21/2022

# -------------------------------------------- file os and other general lib
import os
import shutil
import tempfile
from glob import glob
from tqdm import tqdm
# -------------------------------------------- for data I/O
import openpyxl
import pandas as pd
import zipfile
from inp_parser.inp_parser import Parser
# -------------------------------------------- for command line arguments
import argparse
import sys
# -------------------------------------------- logging
import logging

class batch_curation():
    def __init__(self, base_dir, master_template, mapping_tabular,
            zipped_datafiles, parse_inp, output_zip=None):
        '''
        Input:
        :param base_dir: base directory of the master template, mapping tabular 
            file, zipped datafiles. Output zip file will also be constructed in
            base directory.
        :type base_dir: str
        
        :param master_template: file name of the master template. *.xlsx
        :type master_template: str

        :param mapping_tabular: file name of the mapping tabular file.
            *.xlsx/*.csv/*.tsv
        :type mapping_tabular: str
        
        :param zipped_datafiles: file name of the zipped datafiles. *.zip
        :type zipped_datafiles: str

        :param parse_inp: call inp_parser and modify/overwrite template if True
        :type parse_inp: bool

        :param output_zip: file name of the constructed zip file for batch
            curation. Default to {base_dir}/batch_template_output.zip
        :type output_zip: str or NoneType
        '''
        logging.basicConfig(level=logging.INFO, 
                            filename=f'{base_dir}/batch_curation.log',
                            format='%(asctime)s - %(message)s',
                            datefmt='%d-%b-%y %H:%M:%S')
        # save parse_inp option
        self.parse_inp = parse_inp
        self.base_dir_abspath = os.path.abspath(base_dir)
        self.master_template_abspath = os.path.abspath(os.path.join(base_dir,
            master_template))
        self.zipped_datafiles_abspath = os.path.abspath(os.path.join(base_dir,
            zipped_datafiles))
        # default name of the generated master template in the sample folders
        self.master_template_name = 'master_template.xlsx'
        self.output_zip = os.path.join(self.base_dir_abspath,
            output_zip or 'batch_template_output')
        # trim .zip
        if self.output_zip[-4:] == '.zip':
           self.output_zip = self.output_zip[:-4]
        # load mapping
        self.df = self.read_mapping(self.base_dir_abspath, mapping_tabular)
        # run
        self.run()

    def read_mapping(self, base_dir, mapping_tabular):
        ## Read sample-variable mapping
        file_ext = os.path.splitext(mapping_tabular)[1].lower()
        # Excel file
        if file_ext == '.xlsx':
            df = pd.read_excel(os.path.join(base_dir,mapping_tabular))
        # csv file
        elif file_ext == '.csv':
            df = pd.read_csv(os.path.join(base_dir,mapping_tabular))
        # tsv file
        elif file_ext == '.tsv':
            df = pd.read_csv(os.path.join(base_dir,mapping_tabular),sep='\t')
        else:
            logging.error('The mapping file must be a .xlsx/.csv/.tsv file.')
            raise Exception('The mapping file must be a .xlsx/.csv/.tsv file.')
        return df

    def run(self):
        # create temporary directory
        with tempfile.TemporaryDirectory() as td:
            self.tempdir = td
            # extract zip files
            self.extracted_datafiles_abspath = os.path.join(self.tempdir,
                'zipped_datafiles')
            os.mkdir(self.extracted_datafiles_abspath)
            # move into the temporary directory
            os.chdir(self.tempdir)

            # extract data files
            appendix = zipfile.ZipFile(self.zipped_datafiles_abspath)
            appendix.extractall(self.extracted_datafiles_abspath)
            # for each sample, call run_sample to create the sample folder and move
            # corresponding files into it and modify as needed
            for idx,sample in tqdm(self.df.iterrows(), total=self.df.shape[0]):
                self.run_sample(sample)
            # remove extracted data files
            shutil.rmtree(self.extracted_datafiles_abspath)
            # zip the files in the temporary folder into a zip file self.output_zip
            shutil.make_archive(self.output_zip, 'zip', self.tempdir)
            print(self.output_zip)

    def run_sample(self, sample_mapping):
        logging.info(f'Processing {sample_mapping[0]}.')
        # create a folder for each sample using sample id as the name
        # assume sample id is always the first item in sample_mapping
        sample_folder = os.path.join(self.tempdir,sample_mapping[0])
        os.mkdir(sample_folder)
        # keep track of the inp files
        inps = []
        # copy and paste the appendix datafiles into the folder
        for file in sample_mapping:
            if isinstance(file,str) and \
            os.path.exists(os.path.join(self.extracted_datafiles_abspath,file)):
                shutil.copy(os.path.join(self.extracted_datafiles_abspath,file),
                    os.path.join(sample_folder,file))
                if os.path.splitext(file)[1].lower() == '.inp':
                    inps.append(os.path.join(sample_folder,file))
        # init inp parser output dict
        parser_out = {}
        if self.parse_inp:
            # call parser if .inp file detected, assume only 1 inp file per sample
            if len(inps) > 1:
                print(f'Warning: more than 1 inp files detected for \
{sample_mapping[0]}, only parsing {inps[0]}.')
                logging.warning(f'More than 1 inp files detected for \
{sample_mapping[0]}, only parsing {inps[0]}.')
            if inps:
                # skip *Equation and *Nset
                parser = Parser(inps[0], skip={'*Equation','*Nset'})
                parser_out = parser.to_dict()
        # copy and paste the master template into the folder
        shutil.copy(self.master_template_abspath,
            os.path.join(sample_folder,self.master_template_name))
        # update the master template
        self.update_template(sample_folder, self.master_template_name,
            sample_mapping, parser_out)

    ## Modify master template in place
    def update_template(self, base_dir, master_template, sample_mapping,
        parser_out):
        '''
        Input:
        :param sample_mapping: a row in the mapping DataFrame
        :type sample_mapping: pandas.Series
        '''
        wb = openpyxl.load_workbook(os.path.join(base_dir,master_template))
        ignore_sheets = {'legend','4. Characterization Methods',
            'Characterization Methods to Add','Dropdown menu choices'}
        for sheet in wb:
            # ignore irrelevant worksheets
            if sheet.title in ignore_sheets:
                continue
            # since col 1 is always property names, we can skip it
            for row in sheet.iter_rows(min_col=2):
                for cell in row:
                    if isinstance(cell.value,str) and cell.value.startswith('$'):
                        cell.value = sample_mapping[cell.value]
            # loop through the sheet again to update info parsed from inp files
            for row in sheet.iter_rows():
                if row[0].value in parser_out:
                    row_header = row[0].value
                    # example of parser_out[row_header]:
                    # ['inserted by inp_parser', 'CPE4R']
                    for i,v in enumerate(parser_out[row_header]):
                        if v:
                            if row[i+1].value is not None:
                                print(f'Warning: {row[i+1]} with value {row[i+1].value} overwritten as {v}')
                                logging.warning(f'{row[i+1]} with value {row[i+1].value} overwritten as {v}')
                            row[i+1].value = v
        wb.save(os.path.join(base_dir,master_template))


def readOptions(args=sys.argv[1:]):
    parser = argparse.ArgumentParser(description="Batch master template generation for batch curation into NanoMine.")
    parser.add_argument("-b", "--base_dir", required=True, help="Type the base directory holding your template, mapping, input zip file, and output zip file.")
    parser.add_argument("-t", "--master_template", required=True, help="Type the file name of your filled master template.")
    parser.add_argument("-m", "--mapping_tabular", required=True, help="Type the file name of your mapping tabular file.")
    parser.add_argument("-z", "--zipped_datafiles", required=True, help="Type the file name of the zip file that contains appendix data files.")
    parser.add_argument("-o", "--output_zip", help="[Optional] Type the file name for the output zip file. Default to be 'batch_template_output.zip'")
    parser.add_argument("-p", "--parse_inp", dest='parse_inp', default=False, action='store_true', help="Use this argument to enable inp parsing during batch curation. The corresponding cells in the template will be overwritten.")
    opts = parser.parse_args(args)
    return opts

# Command line operation
if __name__ == '__main__':
    options = readOptions(sys.argv[1:])
    bc = batch_curation(**vars(options))
# Example
# base_dir = './example'
# mapping = 'example_mapping.xlsx'
# template = 'example_template.xlsx'
# zipped_datafiles = 'example.zip'
# bc = batch_curation(base_dir=base_dir, master_template=template,
#     mapping_tabular=mapping,zipped_datafiles=zipped_datafiles)